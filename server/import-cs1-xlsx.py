#!/usr/bin/env python3
"""
Import the private CS1 student workbook into the local SQLite backend.

The workbook contains real student/guardian data, so the .xlsx file is ignored
by Git and must not be committed. This script updates the local database only.
"""

from __future__ import annotations

import json
import os
import re
import sqlite3
import sys
from datetime import date, datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Install it locally with: python -m pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "Danh sach tre CS1.xlsx"
DEFAULT_DB = ROOT / "server" / "data" / "maika.sqlite"
DEFAULT_BACKUP_DIR = ROOT / "server" / "backups"

DEPENDENT_COLLECTIONS = {
    "students",
    "classes",
    "attendance",
    "finance",
    "messages",
    "dailyReports",
    "badges",
}

DEPENDENT_TABLES = [
    "attendance_records",
    "health_records",
    "incidents",
    "invoices",
    "student_consents",
    "notification_reads",
    "notifications",
    "media_asset_students",
    "media_assets",
    "media_albums",
    "tuition_plans",
]

CLASS_COLORS = ["#F97316", "#34D399", "#7C3AED", "#06B6D4", "#EC4899"]


def slugify(value: str) -> str:
    text = value.strip().lower()
    replacements = {
        "đ": "d",
        "ă": "a",
        "â": "a",
        "á": "a",
        "à": "a",
        "ả": "a",
        "ã": "a",
        "ạ": "a",
        "ắ": "a",
        "ằ": "a",
        "ẳ": "a",
        "ẵ": "a",
        "ặ": "a",
        "ấ": "a",
        "ầ": "a",
        "ẩ": "a",
        "ẫ": "a",
        "ậ": "a",
        "ê": "e",
        "é": "e",
        "è": "e",
        "ẻ": "e",
        "ẽ": "e",
        "ẹ": "e",
        "ế": "e",
        "ề": "e",
        "ể": "e",
        "ễ": "e",
        "ệ": "e",
        "í": "i",
        "ì": "i",
        "ỉ": "i",
        "ĩ": "i",
        "ị": "i",
        "ô": "o",
        "ơ": "o",
        "ó": "o",
        "ò": "o",
        "ỏ": "o",
        "õ": "o",
        "ọ": "o",
        "ố": "o",
        "ồ": "o",
        "ổ": "o",
        "ỗ": "o",
        "ộ": "o",
        "ớ": "o",
        "ờ": "o",
        "ở": "o",
        "ỡ": "o",
        "ợ": "o",
        "ư": "u",
        "ú": "u",
        "ù": "u",
        "ủ": "u",
        "ũ": "u",
        "ụ": "u",
        "ứ": "u",
        "ừ": "u",
        "ử": "u",
        "ữ": "u",
        "ự": "u",
        "ý": "y",
        "ỳ": "y",
        "ỷ": "y",
        "ỹ": "y",
        "ỵ": "y",
    }
    for src, dest in replacements.items():
        text = text.replace(src, dest)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "class"


def initials(name: str) -> str:
    words = [w for w in re.split(r"\s+", name.strip()) if w]
    return "".join(w[0].upper() for w in words[-2:])[:3] or "HS"


def cell(value) -> str:
    return str(value).strip() if value is not None else ""


def read_students(workbook_path: Path):
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb.active
    raw_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        values = list(row) + [None] * 8
        stt = cell(values[0])
        student_name = cell(values[1])
        class_name = cell(values[2])
        parent_name = cell(values[3])
        note = cell(values[4])
        if not stt.isdigit() or not student_name or not class_name or not parent_name:
            continue
        raw_rows.append(
            {
                "stt": int(stt),
                "studentName": student_name,
                "className": class_name,
                "parentName": parent_name,
                "note": note,
            }
        )

    seen_classes = []
    for item in raw_rows:
        if item["className"] not in seen_classes:
            seen_classes.append(item["className"])

    classes = []
    class_id_by_name = {}
    for index, name in enumerate(seen_classes, start=1):
        class_id = f"cs1-{slugify(name)}"
        class_id_by_name[name] = class_id
        classes.append(
            {
                "id": class_id,
                "name": name,
                "ageGroup": name,
                "teacherId": None,
                "color": CLASS_COLORS[(index - 1) % len(CLASS_COLORS)],
            }
        )

    enroll_date = date.today().isoformat()
    students = []
    for item in raw_rows:
        student_id = f"cs1-{item['stt']:03d}"
        students.append(
            {
                "id": student_id,
                "name": item["studentName"],
                "dob": "",
                "classId": class_id_by_name[item["className"]],
                "parentName": item["parentName"],
                "parentPhone": "",
                "parentEmail": "",
                "enrollDate": enroll_date,
                "status": "active",
                "initials": initials(item["studentName"]),
                "gender": "unknown",
                "note": item["note"],
                "campus": "CS1",
            }
        )

    return classes, students


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def read_snapshot(conn: sqlite3.Connection):
    rows = conn.execute("SELECT collection, data FROM collection_records ORDER BY collection, id").fetchall()
    snapshot = {}
    for collection, data in rows:
        snapshot.setdefault(collection, []).append(json.loads(data))
    return snapshot


def backup_snapshot(conn: sqlite3.Connection, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    data = read_snapshot(conn)
    created_at = datetime.now(timezone.utc).isoformat()
    payload = {
        "version": 1,
        "app": "maika",
        "reason": "before-cs1-xlsx-import",
        "actor": "server/import-cs1-xlsx.py",
        "createdAt": created_at,
        "collections": {key: len(value) for key, value in data.items()},
        "data": data,
    }
    name = f"maika-backup-before-cs1-import-{created_at.replace(':', '-').replace('.', '-')}.json"
    path = backup_dir / name
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_collection(conn: sqlite3.Connection, collection: str, records):
    conn.execute("DELETE FROM collection_records WHERE collection = ?", (collection,))
    conn.executemany(
        """
        INSERT INTO collection_records (collection, id, data, updated_at)
        VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        """,
        [(collection, record["id"], json.dumps(record, ensure_ascii=False)) for record in records],
    )


def main() -> int:
    workbook = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    db_path = Path(os.environ.get("MAIKA_DB_PATH", DEFAULT_DB)).resolve()
    backup_dir = Path(os.environ.get("MAIKA_BACKUP_DIR", DEFAULT_BACKUP_DIR)).resolve()

    if not workbook.exists():
        print(f"Workbook not found: {workbook}", file=sys.stderr)
        return 1
    if not db_path.exists():
        print(f"Database not found: {db_path}. Run npm run api:dev once to initialize it.", file=sys.stderr)
        return 1

    classes, students = read_students(workbook)
    if not students:
        print("No valid student rows found.", file=sys.stderr)
        return 1

    conn = sqlite3.connect(db_path)
    try:
        if not table_exists(conn, "collection_records"):
            print("Database is missing collection_records. Run npm run api:dev once to initialize it.", file=sys.stderr)
            return 1

        backup_path = backup_snapshot(conn, backup_dir)
        with conn:
            for collection in DEPENDENT_COLLECTIONS:
                conn.execute("DELETE FROM collection_records WHERE collection = ?", (collection,))
            write_collection(conn, "classes", classes)
            write_collection(conn, "students", students)

            for table in DEPENDENT_TABLES:
                if table_exists(conn, table):
                    conn.execute(f"DELETE FROM {table}")

            if table_exists(conn, "users"):
                conn.execute("DELETE FROM users WHERE role = 'parent'")
                parent_rows = [
                    (
                        f"parent-{student['id']}",
                        "parent",
                        student["parentName"] or f"Parent {student['name']}",
                        student["parentPhone"] or None,
                        student["parentEmail"] or None,
                        None,
                        student["id"],
                        "active",
                    )
                    for student in students
                    if student["parentPhone"]
                ]
                parent_account_count = len(parent_rows)
                if parent_rows:
                    conn.executemany(
                        """
                        INSERT OR REPLACE INTO users
                          (id, role, display_name, phone, email, password_hash, student_id, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        parent_rows,
                    )
            else:
                parent_account_count = 0

        print(
            json.dumps(
                {
                    "ok": True,
                    "students": len(students),
                    "classes": len(classes),
                    "parentAccountsCreated": parent_account_count,
                    "backup": str(backup_path.relative_to(ROOT)),
                    "note": "No parent phone column was found in the workbook, so parent login accounts were not created.",
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
