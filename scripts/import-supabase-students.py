#!/usr/bin/env python3
"""Import Maika students from Excel into Supabase.

Required env:
  SUPABASE_URL
  SUPABASE_SERVICE_KEY

Usage:
  python scripts/import-supabase-students.py "Danh sach tre CS1.xlsx" CS1
"""

from __future__ import annotations

import json
import os
import sys
import uuid
import unicodedata
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from urllib.error import HTTPError

from openpyxl import load_workbook


ALIASES = {
    "student": ["họ và tên học sinh", "hoc sinh", "student", "student name", "full_name"],
    "class": ["học lớp", "lop", "class", "class_name"],
    "parent": ["họ và tên cha hoặc mẹ", "phụ huynh", "parent", "parent_name"],
    "phone": ["số điện thoại", "dien thoai", "phone", "parent_phone"],
    "email": ["email", "parent_email"],
    "dob": ["ngày sinh", "dob", "date of birth"],
    "gender": ["giới tính", "gender"],
    "notes": ["ghi chú", "ghi chu", "note", "notes"],
    "facility": ["cơ sở", "co so", "facility", "facility_code"],
}


def norm(value) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def api(path, method="GET", payload=None):
    base = os.environ["SUPABASE_URL"].rstrip("/")
    key = os.environ["SUPABASE_SERVICE_KEY"]
    data = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = Request(
        base + path,
        method=method,
        data=data,
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation,resolution=merge-duplicates",
        },
    )
    try:
        with urlopen(req, timeout=30) as res:
            text = res.read().decode("utf-8")
            return json.loads(text) if text else None
    except HTTPError as exc:
        raise RuntimeError(f"{method} {path} -> {exc.code}: {exc.read().decode('utf-8')}") from exc


def find_header(headers, key):
    normalized = [norm(h) for h in headers]
    for alias in ALIASES[key]:
        alias = norm(alias)
        for index, header in enumerate(normalized):
            if alias == header or alias in header:
                return index
    return None


def facility_id_for(code):
    rows = api("/rest/v1/facilities?" + urlencode({"code": f"eq.{code}", "select": "id"}))
    if not rows:
        raise RuntimeError(f"Facility not found: {code}")
    return rows[0]["id"]


def parse_workbook(path: Path, default_facility_code: str):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_index = next(i for i, row in enumerate(rows) if find_header(row, "student") is not None)
    headers = rows[header_index]
    indexes = {key: find_header(headers, key) for key in ALIASES}
    if indexes["student"] is None:
        raise RuntimeError("Missing student name column")

    facility_cache = {}
    seen = set()
    students = []
    for row in rows[header_index + 1:]:
        get = lambda key: cell_text(row[indexes[key]]) if indexes[key] is not None and indexes[key] < len(row) else ""
        name = get("student")
        parent = get("parent")
        klass = get("class")
        if not name or name.lower().startswith("tổng"):
            continue
        facility_code = get("facility") or default_facility_code
        if facility_code not in facility_cache:
            facility_cache[facility_code] = facility_id_for(facility_code)
        duplicate_key = (facility_code, norm(name), norm(parent), norm(klass))
        if duplicate_key in seen:
            raise RuntimeError(f"Duplicate student row in import file: {facility_code} / {name} / {parent}")
        seen.add(duplicate_key)
        students.append({
            "id": str(uuid.uuid5(uuid.NAMESPACE_URL, f"maika-{facility_code}-{name}-{klass}-{parent}")),
            "facility_id": facility_cache[facility_code],
            "full_name": name,
            "dob": get("dob") or None,
            "gender": {"nam": "male", "nữ": "female", "nu": "female"}.get(norm(get("gender")), "unknown"),
            "class_name": klass or None,
            "parent_name": parent or None,
            "parent_phone": get("phone") or None,
            "parent_email": get("email") or None,
            "status": "active",
            "notes": get("notes") or None,
        })
    return students


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python scripts/import-supabase-students.py workbook.xlsx [CS1|CS2]")
    workbook = Path(sys.argv[1])
    default_facility = sys.argv[2] if len(sys.argv) > 2 else "CS1"
    students = parse_workbook(workbook, default_facility)
    if not students:
        raise SystemExit("No students found")
    result = api("/rest/v1/students?on_conflict=id", method="POST", payload=students)
    print(json.dumps({"ok": True, "studentsImported": len(result or students)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
